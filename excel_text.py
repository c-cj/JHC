import openpyxl
from openpyxl import cell
from openpyxl.styles import Alignment, PatternFill,Border , Side

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

style_set= {'mediumDashDot', 'slantDashDot', 'dotted', 'dashDotDot', 'hair', 'medium', 'dashed',
            'mediumDashDotDot', 'thick', 'double', 'thin', 'dashDot', 'mediumDashed'}
i = int(len(style_set))
r = 1
ws.cell(row=1,column=1).border = Side(style='mediumDashDot')

# for my_style in style_set:
#     ws.cell(row=r,column=1).border = Side(style=my_style)
#     r += 2
wb.save("test.xlsx")